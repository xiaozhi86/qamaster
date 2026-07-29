#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
gen_excel.py — 测试用例 .md → .xlsx 永久生成器（Tier B skill 自带资产）

用途：第15阶段（Excel 生成）由 skill 经 Bash 调用本脚本，读取
`case-design-out/TestCases_<需求标识>.md`，按 `references/excel.md` 21.4 的 15 列字段
规范写出 `case-design-out/TestCases_<需求标识>.xlsx`，并对产出物做结构验证 + 数据
完整性校验（对应 `references/excel.md` 21.7）。

本脚本替代旧"ad-hoc 即兴 openpyxl 脚本 + 用完即删"机制，把 Excel 生成的代码级
单一事实源固化进 skill 自带资产（与 `verify_cases.py`/`verify_md.py` 同级，见
`references/output_write.md` ch30「不清理的 skill 自带资产」清单），消除两类历史缺陷：

1. 中文乱码（根因：ad-hoc 脚本在 Windows 默认 cp936 下 `open(md,"r")` 不带 encoding
   把 UTF-8 中文 .md 按 cp936 错误解码再写进 .xlsx）。
   → 本脚本统一 `open(md_path, "r", encoding="utf-8")`，并 `sys.stdout.reconfigure`。

2. When/Then 不自动换行（根因：ad-hoc 脚本不设 `Alignment(wrap_text=True)`、不设列宽、
   源 .md 的 When/Then 为 `；` 单行无真实换行符）。
   → 本脚本对 Given/When/Then 单元格强制 `Alignment(wrap_text=True, vertical="top")`，
   设宽列 + freeze_panes + auto_filter；写入前按 `；`/`步骤N：` 边界注入真实
   `\\n`(chr(10))，使 wrap_text 有硬换行断点（与 `references/modeling.md` 13.2/13.3
   换行符约定一致）。

脚本须一次执行内完成读取+解析+写出+格式化+两段校验，禁止分次追加写同一 .xlsx
（见 `references/excel.md` 21.7「脚本职责」）。

交付格式硬约束（不可违背）：
  - 测试用例 Excel 交付格式固定为 `.xlsx`（由 openpyxl 的 Workbook 产出）。
  - 禁止以 `.csv` / `.xls` / 其他格式作为测试用例交付物。
  - CSV 仅可作为脚本内部中间产物（若有），用完即删，不得交付给用户。
  - 输出路径缺省 = 源 .md 同名 `.xlsx`（绝不写 `.csv`）。

退出码：0 = 生成成功且两段校验全过；1 = 生成失败/校验不通过（按 21.7「生成失败处理」报错）。
本脚本是 skill 自带可复用资产，不删除（见 `references/output_write.md` ch30）。

用法：
  python gen_excel.py <TC文件.md> [输出.xlsx]
  缺省输出 = 同目录同名 .xlsx（.md→.xlsx，输出到源 .md 所在目录；绝不产出 .csv）。
"""
import sys
import os
import re
import json

# Windows 控制台默认 cp936，强制 stdout 输出 UTF-8，避免核对报告中文乱码
try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

# ===== openpyxl 依赖兜底（references/excel.md 21.7「依赖兜底」）=====
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[Excel 生成失败] 依赖缺失：openpyxl 未安装。正在尝试 pip install openpyxl ...")
    sys.stdout.flush()
    rc = os.system("pip install openpyxl")
    if rc != 0:
        rc = os.system("pip3 install openpyxl")
    if rc != 0:
        print("[Excel 生成失败] openpyxl 自动安装失败。请手动安装后重试：pip install openpyxl")
        sys.exit(1)
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[Excel 生成失败] openpyxl 安装后仍不可用，请检查 Python 环境。")
        sys.exit(1)


# ===== 规则契约单一事实源（与 verify_cases.py / verify_md.py 共同加载 config/validation_rules.json）=====
# 单一事实源：避免表头/枚举/固定值在多份脚本里双份维护、漂移。
def _rules_path():
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "config", "validation_rules.json")


def _load_validation_rules():
    p = _rules_path()
    if not os.path.exists(p):
        print("[Excel 生成失败] 校验规则清单缺失: %s（skill 资产，须与 scripts 同 bundle）" % p)
        return None
    try:
        with open(p, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        print("[Excel 生成失败] 校验规则清单读取失败: %s -> %s" % (p, e))
        return None


_RULES = _load_validation_rules()
if _RULES is None:
    sys.exit(1)

# 15 列顺序（0-based，与 verify_cases.py:74-76 完全一致）
# 0用例ID 1关联需求ID 2关联规则 3测试类型 4测试维度 5所属模块 6用例名称
# 7Given 8When 9Then 10编辑模式 11标签 12责任人 13用例等级 14用例状态
IDX_ID, IDX_REQ, IDX_RULE, IDX_TYPE, IDX_DIM, IDX_MOD, IDX_NAME = 0, 1, 2, 3, 4, 5, 6
IDX_GIVEN, IDX_WHEN, IDX_THEN = 7, 8, 9
IDX_EDITMODE, IDX_TAG, IDX_OWNER, IDX_LEVEL, IDX_STATUS = 10, 11, 12, 13, 14

HEADER_TOKENS = list(_RULES["header"])           # 15 列表头（权威顺序）
EXPECTED_COLS = len(HEADER_TOKENS)                # = 15

VALID_TYPES = set(_RULES["valid_types"])
VALID_DIMS = set(_RULES["valid_dims"])
VALID_LEVELS = set(_RULES["valid_levels"])
_fc = _RULES["fixed_columns"]
FIX_EDITMODE, FIX_TAG, FIX_OWNER, FIX_STATUS = _fc["edit_mode"], _fc["tag"], _fc["owner"], _fc["status"]

NAME_SEGMENTS = _RULES["name_segments"]
VAGUE_WORDS = list(_RULES["vague_words"])
OBSERVABLE_PATTERNS = list(_RULES["observable_patterns"])
STORAGE_PATTERNS = [(p["pattern"], p["desc"]) for p in _RULES["storage_patterns"]]
STORAGE_NATURAL = list(_RULES["storage_natural"])

# 需填真实换行符的长文本列（Given/When/Then）
WRAP_COLS = (IDX_GIVEN, IDX_WHEN, IDX_THEN)
# 列宽（字符数）；Given/When/Then 宽列，其余按内容估宽的固定值（openpyxl 无真自动列宽）
COL_WIDTHS = {
    IDX_ID: 22,
    IDX_REQ: 18,
    IDX_RULE: 18,
    IDX_TYPE: 10,
    IDX_DIM: 12,
    IDX_MOD: 12,
    IDX_NAME: 34,
    IDX_GIVEN: 50,
    IDX_WHEN: 55,
    IDX_THEN: 60,
    IDX_EDITMODE: 10,
    IDX_TAG: 8,
    IDX_OWNER: 8,
    IDX_LEVEL: 8,
    IDX_STATUS: 12,
}


# ===== .md 用例表解析（与 verify_cases.py:190-248 共享 split_row/is_separator/parse_table 逻辑）=====
def split_row(line):
    s = line.strip()
    if not s.startswith("|"):
        return None
    s = s[1:]
    if s.endswith("|"):
        s = s[:-1]
    return [c.strip() for c in s.split("|")]


def is_separator(cells):
    if not cells:
        return False
    for c in cells:
        if not set(c) <= set("-: "):
            return False
    return True


def parse_table(path):
    """解析 .md 中的用例表，返回 (header_cells, data_rows) 或 (None, err)。

    以"用例ID"表头行定位用例表起始；表前的"规则建模/风险清单/测试点清单"等追溯性
    section 不读取、不校验、不输出到 Excel（见 references/excel.md 21.7「脚本职责」）。
    """
    if not os.path.exists(path):
        return None, "文件不存在: %s" % path
    try:
        with open(path, "r", encoding="utf-8") as f:  # 强制 UTF-8，堵中文乱码根因
            lines = f.readlines()
    except Exception as e:
        return None, "读取失败: %s" % e

    header_idx = None
    header_cells = None
    for i, ln in enumerate(lines):
        cells = split_row(ln)
        if cells and "用例ID" in cells[0]:
            header_idx = i
            header_cells = cells
            break
    if header_idx is None:
        return None, "未找到表头行（含'用例ID'的表格行）"

    data_rows = []
    for ln in lines[header_idx + 1:]:
        cells = split_row(ln)
        if cells is None:
            if data_rows:
                break
            else:
                continue
        if is_separator(cells) or len(cells) == 0:
            continue
        data_rows.append(cells)
    return (header_cells, data_rows), None


def count_segments(name):
    return len(re.findall(r"【[^】]*】", name))


# ===== When/Then 换行符注入（堵自动换行根因第3点）=====
# 源 .md 的 When/Then 在 Markdown 单行内常以 `；` 或 `步骤N：` 串联；Excel 单元格需
# 真实 \n(chr(10)) 才能让 wrap_text 在步骤边界硬换行。这里把边界替换为 \n，保留步骤序号。
_STEP_BOUNDARY = re.compile(r"(？需{0}|步骤\s*\d+\s*[:：]|(?<=[^；\n])；(?=\s*(?:步骤|\d+\.)))")


def inject_newlines(text):
    """把 When/Then/Given 单行 `；`/`步骤N：` 边界注入为真实 \\n，给 wrap_text 提供硬换行断点。

    规则：
    - `步骤N：` 前注入换行（每步一行）；
    - `数字.` 编号前若无换行则注入；
    - 纯 `；` 分隔的步骤也注入换行（每步一行），不依赖 `步骤N：`/`N.` 标记；
    - 已含真实 \\n/\\r 的保留。
    """
    if not text:
        return text
    # 统一换行符为 \n
    t = text.replace("\r\n", "\n").replace("\r", "\n")
    # 在 `步骤N：` 前插入换行（若不在行首）
    t = re.sub(r"(?<!^)(?<!\n)(\s*步骤\s*\d+\s*[:：])", r"\n\1", t)
    # 在 `1.`/`2.` 编号前插入换行（若不在行首且未在换行后）—— 形如 "1.xxx；2.yyy"
    t = re.sub(r"(?<!^)(?<!\n)(?<=；)\s*(\d+\.\s)", r"\n\1", t)
    # 纯 `；` 分隔也注入换行（每步一行），使只用 `；` 分隔时 Excel 也能按步骤硬换行
    t = re.sub(r"(?<!^)(?<!\n)\s*；\s*", "\n", t)
    return t.strip()


# ===== 校验：数据完整性九项（对应 references/excel.md 21.7「数据完整性校验」，复用 verify_cases.py 口径）=====
def check_cell_consistency(data_rows):
    """1. 逐单元格一致性：表头列数须=15、每行列数=15。返回违规列表。"""
    v = []
    for i, r in enumerate(data_rows, 1):
        if len(r) != EXPECTED_COLS:
            v.append("行%d: 列数=%d（应为15）" % (i, len(r)))
    return v


def check_required_nonempty(data_rows):
    """2. 必填字段非空：用例ID/关联需求ID/关联规则/测试类型/所属模块/测试维度/用例名称/Given/When/Then。"""
    required = [IDX_ID, IDX_REQ, IDX_RULE, IDX_TYPE, IDX_MOD, IDX_DIM, IDX_NAME, IDX_GIVEN, IDX_WHEN, IDX_THEN]
    v = []
    for i, r in enumerate(data_rows, 1):
        for idx in required:
            val = r[idx].strip() if idx < len(r) else ""
            if not val:
                v.append("行%d: 必填字段『%s』为空" % (i, HEADER_TOKENS[idx]))
    return v


def check_fixed_columns(data_rows):
    """3. 固定列取值正确：编辑模式=STEP、标签=AI、责任人=AI、用例状态=Completed。"""
    v = []
    for i, r in enumerate(data_rows, 1):
        if r[IDX_EDITMODE] != FIX_EDITMODE:
            v.append("行%d: 编辑模式应为%s，实际『%s』" % (i, FIX_EDITMODE, r[IDX_EDITMODE]))
        if r[IDX_TAG] != FIX_TAG:
            v.append("行%d: 标签应为%s，实际『%s』" % (i, FIX_TAG, r[IDX_TAG]))
        if r[IDX_OWNER] != FIX_OWNER:
            v.append("行%d: 责任人应为%s，实际『%s』" % (i, FIX_OWNER, r[IDX_OWNER]))
        if r[IDX_STATUS] != FIX_STATUS:
            v.append("行%d: 用例状态应为%s，实际『%s』" % (i, FIX_STATUS, r[IDX_STATUS]))
    return v


def check_level(data_rows):
    """4. 用例等级合规：P0-P3。"""
    v = []
    for i, r in enumerate(data_rows, 1):
        if r[IDX_LEVEL] not in VALID_LEVELS:
            v.append("行%d: 用例等级『%s』越界（允许P0-P3）" % (i, r[IDX_LEVEL]))
    return v


def check_ids(data_rows):
    """5. ID 唯一不跳号：用例ID 全局唯一、按功能缩写分组连续。"""
    v = []
    seen = {}
    for i, r in enumerate(data_rows, 1):
        cid = r[IDX_ID]
        if cid in seen:
            v.append("行%d: 用例ID重复『%s』（首次出现于行%d）" % (i, cid, seen[cid]))
        else:
            seen[cid] = i
    # 跳号：按功能缩写分组
    groups = {}
    for i, r in enumerate(data_rows, 1):
        parts = r[IDX_ID].split("_")
        if len(parts) < 2:
            continue
        seq = parts[-1]
        func = parts[-2] if len(parts) >= 3 else None
        if func is None:
            continue
        try:
            n = int(seq)
        except Exception:
            continue
        groups.setdefault(func, []).append(n)
    for func, nums in groups.items():
        nums = sorted(nums)
        if not nums:
            continue
        expected = list(range(nums[0], nums[-1] + 1))
        missing = set(expected) - set(nums)
        if missing:
            v.append("功能[%s]序号跳号，缺失: %s" % (func, ",".join(str(x) for x in sorted(missing))))
    return v


def check_enums(data_rows):
    """6. 枚举合规：测试类型/测试维度取值在允许枚举内。"""
    v = []
    for i, r in enumerate(data_rows, 1):
        if r[IDX_TYPE] not in VALID_TYPES:
            v.append("行%d: 测试类型『%s』越界" % (i, r[IDX_TYPE]))
        if r[IDX_DIM] not in VALID_DIMS:
            v.append("行%d: 测试维度『%s』越界" % (i, r[IDX_DIM]))
    return v


def check_name_spec(data_rows):
    """7. 用例名称规范：四段【模块】【功能】【场景】【预期】。"""
    v = []
    for i, r in enumerate(data_rows, 1):
        if count_segments(r[IDX_NAME]) < NAME_SEGMENTS:
            v.append("行%d: 用例名称段数不足（%d<%d）『%s』" % (i, count_segments(r[IDX_NAME]), NAME_SEGMENTS, r[IDX_NAME]))
    return v


def check_assertions(data_rows):
    """8. 断言可观测：Then 含可观测关键词、不含模糊词（软判定，列疑似条数）。"""
    suspects = []
    for i, r in enumerate(data_rows, 1):
        then = r[IDX_THEN]
        if not then.strip():
            suspects.append("行%d: Then 为空" % i)
            continue
        if any(re.search(p, then) for p in OBSERVABLE_PATTERNS):
            continue
        vague_hit = [w for w in VAGUE_WORDS if w in then]
        if vague_hit:
            suspects.append("行%d: Then 疑似模糊断言（含%s且无可观测锚点）" % (i, "/".join(vague_hit)))
        else:
            suspects.append("行%d: Then 未识别到可观测锚点（请人工复核）" % i)
    return len(suspects), suspects


def check_storage(data_rows):
    """9. 存储合规：无杜撰表名/字段名/Redis Key/Topic/Index/Bucket（软判定，列疑似条数）。"""
    suspects = []
    for i, r in enumerate(data_rows, 1):
        text = " ".join(r[IDX_GIVEN:IDX_THEN + 1]) if len(r) > IDX_THEN else ""
        hits = [desc for pat, desc in STORAGE_PATTERNS if re.search(pat, text, flags=re.IGNORECASE)]
        if hits:
            natural = [n for n in STORAGE_NATURAL if n in text]
            tag = "（含自然语言描述，请复核）" if natural else ""
            suspects.append("行%d: %s%s" % (i, ";".join(hits), tag))
    return len(suspects), suspects


# ===== 生成 .xlsx =====
def build_xlsx(header_cells, data_rows, out_path):
    """按 21.4 字段顺序写出 .xlsx：表头行 + 逐条用例行 + 格式化。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "TestCases"

    # --- 表头行（第1行）---
    bold_center_wrap = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c, title in enumerate(HEADER_TOKENS, 1):
        cell = ws.cell(row=1, column=c, value=title)
        cell.font = bold_center_wrap
        cell.alignment = header_align

    # --- 数据行（第2行起）---
    top_wrap = Alignment(wrap_text=True, vertical="top")
    top_left = Alignment(wrap_text=True, vertical="top", horizontal="left")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ri, row in enumerate(data_rows, 2):
        # 容错：行列数不足 15 补空，超过 15 截断（校验会报违规）
        cells = list(row) + [""] * (EXPECTED_COLS - len(row)) if len(row) < EXPECTED_COLS else list(row[:EXPECTED_COLS])
        for ci, val in enumerate(cells, 1):
            idx = ci - 1
            # Given/When/Then：注入真实换行符，使 wrap_text 在步骤边界硬换行
            if idx in WRAP_COLS:
                val = inject_newlines(val)
            cell = ws.cell(row=ri, column=ci, value=val)
            if idx in WRAP_COLS:
                cell.alignment = top_wrap if idx == IDX_THEN else top_left
            elif idx in (IDX_ID, IDX_TYPE, IDX_DIM, IDX_MOD, IDX_LEVEL, IDX_STATUS,
                         IDX_EDITMODE, IDX_TAG, IDX_OWNER):
                cell.alignment = center
            else:
                cell.alignment = top_left

    # --- 列宽（openpyxl 无真自动列宽，显式设宽）---
    for idx, w in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(idx + 1)].width = w

    # --- 表头冻结 + 自动筛选（references/excel.md 21.2）---
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # --- 行高：openpyxl 不设行高时，多数 Excel 查看器对 wrap_text=True 会按内容自动撑高；
    #     但部分查看器不重算。对 Given/When/Then 任一含 \n 的行，按 \n 数兜底设一个最小高度，
    #     避免换行文本被垂直隐藏（堵自动换行根因第4点）。---
    for ri, row in enumerate(data_rows, 2):
        cells = list(row) + [""] * (EXPECTED_COLS - len(row)) if len(row) < EXPECTED_COLS else list(row[:EXPECTED_COLS])
        max_lines = 1
        for idx in WRAP_COLS:
            v = inject_newlines(cells[idx]) if cells[idx] else ""
            # 估算显示行数：换行数 + 按列宽折行的行数
            lines = v.split("\n") if v else [""]
            wrapped = 0
            for ln in lines:
                wl = max(1, -(-len(ln) // max(8, int(COL_WIDTHS[idx] * 0.9))))  # 向上取整
                wrapped += wl
            max_lines = max(max_lines, wrapped)
        if max_lines > 1:
            # 每行约 15 磅，最小 30
            ws.row_dimensions[ri].height = max(30, min(max_lines * 15, 409))

    wb.save(out_path)


# ===== 生成后两段校验（references/excel.md 21.7「生成后强制验证」+「数据完整性校验」）=====
def verify_structure(out_path, expected_rows):
    """结构验证：落盘非空 / openpyxl 可读 / 数据行数=N / 列数=15 且表头顺序一致。返回 (ok, details)。"""
    details = []
    if not os.path.exists(out_path):
        return False, ["落盘校验: .xlsx 不存在"]
    if os.path.getsize(out_path) <= 0:
        return False, ["落盘校验: .xlsx 为空"]
    details.append("落盘校验: 通过（大小=%d 字节）" % os.path.getsize(out_path))
    try:
        wb = load_workbook(out_path, read_only=True)
        ws = wb.active
    except Exception as e:
        return False, ["可读校验: openpyxl 无法打开 -> %s" % e]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return False, ["可读校验: 表为空"]
    header = [str(c) if c is not None else "" for c in rows[0]]
    if len(header) != EXPECTED_COLS:
        wb.close()
        return False, ["列数校验: 表头列数=%d（应为15）" % len(header)]
    if header != HEADER_TOKENS:
        wb.close()
        return False, ["列数校验: 表头顺序与 21.4 不一致\n  实际: %s\n  期望: %s" % (header, HEADER_TOKENS)]
    details.append("列数校验: 通过（15列，表头顺序与21.4一致）")
    data_n = len(rows) - 1
    if data_n != expected_rows:
        wb.close()
        return False, ["行数校验: 数据行=%d（源.md用例数=%d）" % (data_n, expected_rows)]
    details.append("行数校验: 通过（数据行=%d）" % data_n)
    wb.close()
    return True, details


def verify_data_integrity(data_rows):
    """数据完整性九项（对应 references/excel.md 21.7），返回 (硬违规列表, 软疑似统计)。"""
    hard = []
    soft = []

    v = check_cell_consistency(data_rows)
    if v:
        hard.append(("逐单元格一致", v))
    v = check_required_nonempty(data_rows)
    if v:
        hard.append(("必填非空", v))
    v = check_fixed_columns(data_rows)
    if v:
        hard.append(("固定列取值", v))
    v = check_level(data_rows)
    if v:
        hard.append(("用例等级合规", v))
    v = check_ids(data_rows)
    if v:
        hard.append(("ID唯一不跳号", v))
    v = check_enums(data_rows)
    if v:
        hard.append(("枚举合规", v))
    v = check_name_spec(data_rows)
    if v:
        hard.append(("用例名称规范", v))

    n_assert, suspect_assert = check_assertions(data_rows)
    n_store, suspect_store = check_storage(data_rows)
    soft.append(("断言可观测", n_assert, suspect_assert))
    soft.append(("存储合规", n_store, suspect_store))

    return hard, soft


# ===== 主入口 =====
def main():
    if len(sys.argv) < 2:
        print("用法: python gen_excel.py <TC文件.md> [输出.xlsx]")
        return 1

    src_md = sys.argv[1]
    if len(sys.argv) >= 3:
        out_path = sys.argv[2]
    else:
        # 默认同名 .md→.xlsx，输出到源 .md 所在目录
        base, _ = os.path.splitext(src_md)
        out_path = base + ".xlsx"

    # 1. 解析源 .md
    parsed, err = parse_table(src_md)
    if parsed is None:
        print("[Excel 生成失败] 解析失败: %s" % err)
        return 1
    header_cells, data_rows = parsed

    # 表头顺序一致性（转换前核对）
    if header_cells[:EXPECTED_COLS] != HEADER_TOKENS:
        print("[Excel 生成失败] 源 .md 表头顺序与 21.4 不一致：")
        print("  实际: %s" % header_cells[:EXPECTED_COLS])
        print("  期望: %s" % HEADER_TOKENS)
        print("请先修正源 .md 表头（见 references/excel.md 21.6 一致性校验）。")
        return 1

    n = len(data_rows)
    if n == 0:
        print("[Excel 生成失败] 源 .md 用例数为 0，无可转换内容。")
        return 1

    # 2. 生成 .xlsx
    try:
        build_xlsx(header_cells, data_rows, out_path)
    except Exception as e:
        print("[Excel 生成失败] 脚本异常: %s" % e)
        return 1

    # 3. 生成后强制验证（结构四项）
    ok, details = verify_structure(out_path, n)
    print("===== Excel 结构验证 =====")
    print("文件: %s" % os.path.basename(out_path))
    for d in details:
        print(d)
    if not ok:
        print("结构验证: 不通过 -> 生成失败，禁止交付")
        print("========================")
        return 1
    print("结构验证: 通过")
    print("========================")

    # 4. 数据完整性校验（九项）
    hard, soft = verify_data_integrity(data_rows)
    print("===== Excel 数据完整性核对报告 =====")
    print("| 校验项 | 核对结果 | 依据 |")
    print("| -- | ---- | ---- |")

    def fmt_hard(name, v):
        return ("不通过", "违规条数=%d（%s）" % (len(v), "; ".join(v[:5]) + ("..." if len(v) > 5 else "")))

    for name, v in hard:
        res, basis = fmt_hard(name, v)
        print("| %s | %s | %s |" % (name, res, basis))
    # 未违规的硬性项也列出（通过）
    hard_names = [name for name, _ in hard]
    for name in ["逐单元格一致", "必填非空", "固定列取值", "用例等级合规", "ID唯一不跳号", "枚举合规", "用例名称规范"]:
        if name not in hard_names:
            print("| %s | 通过 | 违规条数=0 |" % name)

    for name, cnt, suspects in soft:
        if cnt == 0:
            print("| %s | 通过 | 疑似条数=0 |" % name)
        else:
            sample = "; ".join(suspects[:3]) + ("..." if len(suspects) > 3 else "")
            print("| %s | 通过(软判定) | 疑似条数=%d（%s）" % (name, cnt, sample))

    print("========================")

    if hard:
        print("[Excel 生成失败] 数据完整性校验存在硬性违规（%d 项），详见上表。" % len(hard))
        print("按 references/excel.md 21.7「校验不通过处理」：结构类问题须修正脚本逻辑或源 .md 后重新整体生成 .xlsx，禁止用 Edit 补打 Excel 单元格。")
        return 1

    # 5. 成功
    print("[Excel 生成成功] %s（%d 条用例，结构验证 + 数据完整性校验全过）" % (out_path, n))
    print("格式特性：UTF-8 兼容 / Given·When·Then 自动换行(wrap_text=True) / 列宽已设 / 表头冻结 / 支持筛选 / 一行一用例 / 无合并单元格")
    return 0


if __name__ == "__main__":
    sys.exit(main())
